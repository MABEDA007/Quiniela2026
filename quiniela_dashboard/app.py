from __future__ import annotations

import argparse
import base64
import binascii
import csv
import datetime as dt
import hmac
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl


APP_DIR = Path(__file__).resolve().parent
DEFAULT_QUINIELA_DIR = APP_DIR / "data"
CACHE_DIR = Path(os.environ.get("QUINIELA_CACHE_DIR", APP_DIR / "cache"))
STATIC_DIR = APP_DIR / "static"
SEASON = 2026
ESPN_SCOREBOARD_URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"
ESPN_DATES = "20260611-20260719"
LOCAL_TZ = ZoneInfo(os.environ.get("QUINIELA_TIMEZONE", "America/Mexico_City"))
MAX_UPLOAD_BYTES = 12 * 1024 * 1024
CSV_FIELDS = [
    "participant",
    "file",
    "sheet",
    "matchNo",
    "date",
    "time",
    "home",
    "predHome",
    "predAway",
    "away",
    "venue",
]

COMPLETED_STATUSES = {"FT", "AET", "PEN", "STATUS_FINAL", "STATUS_FULL_TIME", "STATUS_FINAL_PEN"}
NOT_PLAYED_STATUSES = {"TBD", "NS", "PST", "CANC", "ABD", "AWD", "WO", "STATUS_SCHEDULED", "STATUS_POSTPONED", "STATUS_CANCELED"}

TEAM_ALIASES = {
    "alemania": "germany",
    "arabia saudita": "saudi arabia",
    "argelia": "algeria",
    "australia": "australia",
    "austria": "austria",
    "belgica": "belgium",
    "belgium": "belgium",
    "bosnia y herzegovina": "bosnia and herzegovina",
    "bosnia-herzegovina": "bosnia and herzegovina",
    "brasil": "brazil",
    "brazil": "brazil",
    "cabo verde": "cape verde",
    "canada": "canada",
    "catar": "qatar",
    "colombia": "colombia",
    "corea del sur": "south korea",
    "korea republic": "south korea",
    "south korea": "south korea",
    "costa de marfil": "ivory coast",
    "cote divoire": "ivory coast",
    "cote d ivoire": "ivory coast",
    "croacia": "croatia",
    "curacao": "curacao",
    "ecuador": "ecuador",
    "egipto": "egypt",
    "escocia": "scotland",
    "espana": "spain",
    "estados unidos": "united states",
    "usa": "united states",
    "francia": "france",
    "ghana": "ghana",
    "haiti": "haiti",
    "inglaterra": "england",
    "iran": "iran",
    "irak": "iraq",
    "iraq": "iraq",
    "japon": "japan",
    "jordania": "jordan",
    "marruecos": "morocco",
    "mexico": "mexico",
    "noruega": "norway",
    "nueva zelanda": "new zealand",
    "paises bajos": "netherlands",
    "panama": "panama",
    "paraguay": "paraguay",
    "portugal": "portugal",
    "rd congo": "congo dr",
    "congo dr": "congo dr",
    "republica checa": "czech republic",
    "czechia": "czech republic",
    "senegal": "senegal",
    "sudafrica": "south africa",
    "suecia": "sweden",
    "suiza": "switzerland",
    "tunez": "tunisia",
    "turquia": "turkey",
    "uruguay": "uruguay",
    "uzbekistan": "uzbekistan",
}

LOG_HANDLE = None


def ensure_output_streams() -> None:
    global LOG_HANDLE
    if sys.stdout is not None and sys.stderr is not None:
        return
    LOG_HANDLE = (APP_DIR / "server.log").open("a", encoding="utf-8")
    if sys.stdout is None:
        sys.stdout = LOG_HANDLE
    if sys.stderr is None:
        sys.stderr = LOG_HANDLE


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def team_key(name: Any) -> str:
    normalized = normalize_text(name)
    return TEAM_ALIASES.get(normalized, normalized)


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def excel_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        match = re.search(r"\d{4}-\d{2}-\d{2}", value)
        if match:
            return match.group(0)
    return None


def excel_time(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, str):
        match = re.search(r"\d{1,2}:\d{2}", value)
        if match:
            return match.group(0).zfill(5)
    return None


def participant_from_workbook(path: Path, worksheet: Any) -> str:
    for row in range(1, 6):
        for col in range(1, min(worksheet.max_column, 16) + 1):
            value = worksheet.cell(row, col).value
            if isinstance(value, str) and normalize_text(value) in {"participante", "jugador"}:
                for next_col in range(col + 1, min(col + 6, worksheet.max_column) + 1):
                    candidate = worksheet.cell(row, next_col).value
                    if isinstance(candidate, str) and normalize_text(candidate) in {"total puntos", "puntos", "pts"}:
                        break
                    if candidate:
                        return str(candidate).strip()

    stem = path.stem
    stem = re.sub(r"(?i)^quiniela[_\s-]*copa[_\s-]*mundial[_\s-]*fifa[_\s-]*2026[_\s-]*", "", stem)
    stem = re.sub(r"(?i)jornada\s*\d+", "", stem)
    stem = re.sub(r"[_-]+", " ", stem).strip()
    return stem or path.stem


def find_prediction_rows(worksheet: Any) -> list[int]:
    rows: list[int] = []
    for row in range(1, worksheet.max_row + 1):
        number = worksheet.cell(row, 5).value
        home = worksheet.cell(row, 9).value
        home_goals = worksheet.cell(row, 10).value
        away_goals = worksheet.cell(row, 12).value
        away = worksheet.cell(row, 13).value
        if as_int(number) is not None and home and away and as_int(home_goals) is not None and as_int(away_goals) is not None:
            rows.append(row)
    return rows


def read_quinielas(quiniela_dir: Path) -> list[dict[str, Any]]:
    participants: list[dict[str, Any]] = []
    quiniela_dir.mkdir(parents=True, exist_ok=True)
    participants.extend(read_csv_quinielas(quiniela_dir))
    files = sorted(path for path in quiniela_dir.glob("*.xlsx") if not path.name.startswith("~$"))

    for path in files:
        try:
            workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
        except Exception as exc:
            participants.append(
                {
                    "name": path.stem,
                    "file": path.name,
                    "error": f"No pude leer el archivo: {exc}",
                    "predictions": [],
                }
            )
            continue

        for worksheet in workbook.worksheets:
            rows = find_prediction_rows(worksheet)
            if not rows:
                continue

            predictions = []
            for row in rows:
                home = worksheet.cell(row, 9).value
                away = worksheet.cell(row, 13).value
                predictions.append(
                    {
                        "matchNo": as_int(worksheet.cell(row, 5).value),
                        "date": excel_date(worksheet.cell(row, 6).value),
                        "time": excel_time(worksheet.cell(row, 7).value),
                        "home": str(home).strip(),
                        "away": str(away).strip(),
                        "homeKey": team_key(home),
                        "awayKey": team_key(away),
                        "predHome": as_int(worksheet.cell(row, 10).value),
                        "predAway": as_int(worksheet.cell(row, 12).value),
                        "venue": worksheet.cell(row, 15).value,
                        "sheet": worksheet.title,
                    }
                )

            name = participant_from_workbook(path, worksheet)
            if len(workbook.worksheets) > 1 and worksheet.title.lower() != "quiniela":
                name = f"{name} ({worksheet.title})"

            participants.append({"name": name, "file": path.name, "predictions": predictions})

    return participants


def read_csv_quinielas(quiniela_dir: Path) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    files = sorted(quiniela_dir.glob("*.csv"))

    for path in files:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                participant = str(row.get("participant") or path.stem).strip()
                source_file = str(row.get("file") or path.name).strip()
                group_key = (participant, source_file)
                group = grouped.setdefault(
                    group_key,
                    {"name": participant, "file": source_file, "predictions": []},
                )

                home = row.get("home")
                away = row.get("away")
                if not home or not away:
                    continue

                group["predictions"].append(
                    {
                        "matchNo": as_int(row.get("matchNo")),
                        "date": row.get("date") or None,
                        "time": row.get("time") or None,
                        "home": str(home).strip(),
                        "away": str(away).strip(),
                        "homeKey": team_key(home),
                        "awayKey": team_key(away),
                        "predHome": as_int(row.get("predHome")),
                        "predAway": as_int(row.get("predAway")),
                        "venue": row.get("venue") or None,
                        "sheet": row.get("sheet") or path.stem,
                    }
                )

    return list(grouped.values())


def cache_path(name: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return CACHE_DIR / name


def read_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def api_get(url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    query = urllib.parse.urlencode(params or {})
    api_url = f"{url}?{query}" if query else url
    request = urllib.request.Request(api_url, headers={"Accept": "application/json", "User-Agent": "quiniela-dashboard/1.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def parse_espn_date(value: Any) -> tuple[str | None, str | None]:
    if not isinstance(value, str):
        return None, None
    try:
        raw = value.replace("Z", "+00:00")
        date_value = dt.datetime.fromisoformat(raw)
        local_value = date_value.astimezone(LOCAL_TZ)
        return date_value.isoformat().replace("+00:00", "Z"), local_value.date().isoformat()
    except ValueError:
        return value, value[:10]


def espn_competitor(competition: dict[str, Any], home_away: str) -> dict[str, Any]:
    for competitor in competition.get("competitors", []):
        if competitor.get("homeAway") == home_away:
            return competitor
    return {}


def normalize_espn_fixture(raw: dict[str, Any]) -> dict[str, Any]:
    competition = (raw.get("competitions") or [{}])[0]
    status = raw.get("status") or competition.get("status") or {}
    status_type = status.get("type") or {}
    venue = competition.get("venue") or raw.get("venue") or {}
    venue_address = venue.get("address") or {}
    home = espn_competitor(competition, "home")
    away = espn_competitor(competition, "away")
    home_team = home.get("team") or {}
    away_team = away.get("team") or {}
    date_value, local_date = parse_espn_date(raw.get("date") or competition.get("date"))

    return {
        "id": raw.get("id"),
        "date": date_value,
        "localDate": local_date,
        "round": (raw.get("season") or {}).get("slug"),
        "leagueLogo": None,
        "venue": venue.get("fullName") or venue.get("displayName"),
        "city": venue_address.get("city"),
        "statusShort": status_type.get("shortDetail") or status_type.get("detail") or status_type.get("name"),
        "statusLong": status_type.get("description") or status_type.get("name"),
        "statusState": status_type.get("state"),
        "completed": bool(status_type.get("completed")),
        "elapsed": status.get("period"),
        "home": home_team.get("displayName") or home_team.get("name"),
        "away": away_team.get("displayName") or away_team.get("name"),
        "homeLogo": home_team.get("logo"),
        "awayLogo": away_team.get("logo"),
        "homeKey": team_key(home_team.get("displayName") or home_team.get("name")),
        "awayKey": team_key(away_team.get("displayName") or away_team.get("name")),
        "homeGoals": as_int(home.get("score")),
        "awayGoals": as_int(away.get("score")),
    }


def load_fixtures(refresh: bool = False) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = cache_path("fixtures_2026.json")
    cached = read_json(path)
    cache_age = None
    if path.exists():
        cache_age = int(dt.datetime.now().timestamp() - path.stat().st_mtime)

    should_use_cache = cached and not refresh and cache_age is not None and cache_age < 900
    if should_use_cache:
        return cached.get("fixtures", []), cached.get("meta", {})

    try:
        data = api_get(ESPN_SCOREBOARD_URL, {"dates": ESPN_DATES, "limit": 200})
        fixtures = [normalize_espn_fixture(item) for item in data.get("events", [])]
        meta = {
            "source": "espn",
            "fetchedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
            "count": len(fixtures),
            "endpoint": ESPN_SCOREBOARD_URL,
            "dates": ESPN_DATES,
        }
        path.write_text(json.dumps({"fixtures": fixtures, "meta": meta}, ensure_ascii=False, indent=2), encoding="utf-8")
        return fixtures, meta
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        if cached:
            meta = cached.get("meta", {})
            meta["source"] = "cache"
            meta["warning"] = f"No pude consultar ESPN; estoy usando cache local. Detalle: {exc}"
            return cached.get("fixtures", []), meta
        return [], {"source": "error", "warning": f"No pude consultar ESPN: {exc}"}


def match_fixture(prediction: dict[str, Any], fixtures: list[dict[str, Any]]) -> dict[str, Any] | None:
    best_fixture = None
    best_score = 0
    for fixture in fixtures:
        same_order = prediction["homeKey"] == fixture["homeKey"] and prediction["awayKey"] == fixture["awayKey"]
        reverse_order = prediction["homeKey"] == fixture["awayKey"] and prediction["awayKey"] == fixture["homeKey"]
        if not same_order and not reverse_order:
            continue

        score = 90 if same_order else 75
        if prediction.get("date") and prediction["date"] == fixture.get("localDate"):
            score += 15
        if prediction.get("venue") and fixture.get("venue") and normalize_text(prediction["venue"]) in normalize_text(fixture["venue"]):
            score += 5
        if score > best_score:
            best_score = score
            best_fixture = fixture

    return best_fixture if best_score >= 75 else None


def result_sign(home_goals: int, away_goals: int) -> int:
    if home_goals > away_goals:
        return 1
    if home_goals < away_goals:
        return -1
    return 0


def score_prediction(prediction: dict[str, Any], fixture: dict[str, Any] | None) -> dict[str, Any]:
    if not fixture:
        return {"points": None, "status": "unmatched", "label": "Sin vincular"}

    actual_home = as_int(fixture.get("homeGoals"))
    actual_away = as_int(fixture.get("awayGoals"))
    status = fixture.get("statusShort")
    state = fixture.get("statusState")
    has_playable_score = state == "in" or bool(fixture.get("completed")) or status in COMPLETED_STATUSES
    if not has_playable_score or status in NOT_PLAYED_STATUSES or actual_home is None or actual_away is None:
        return {"points": None, "status": "pending", "label": "Pendiente"}

    pred_home = prediction["predHome"]
    pred_away = prediction["predAway"]
    if pred_home == actual_home and pred_away == actual_away:
        return {"points": 3, "status": "exact", "label": "Exacto"}
    if result_sign(pred_home, pred_away) == result_sign(actual_home, actual_away):
        return {"points": 1, "status": "trend", "label": "Tendencia"}
    return {"points": 0, "status": "miss", "label": "Fallo"}


def build_state(quiniela_dir: Path, refresh: bool = False) -> dict[str, Any]:
    participants = read_quinielas(quiniela_dir)
    fixtures, api_meta = load_fixtures(refresh=refresh)
    enriched_participants = []
    all_predictions = []

    for participant in participants:
        stats = {"points": 0, "exact": 0, "trend": 0, "miss": 0, "pending": 0, "unmatched": 0}
        predictions = []
        for prediction in participant.get("predictions", []):
            fixture = match_fixture(prediction, fixtures)
            score = score_prediction(prediction, fixture)
            if score["points"] is None:
                stats[score["status"]] += 1
            else:
                stats["points"] += score["points"]
                stats[score["status"]] += 1

            enriched = {**prediction, "fixture": fixture, "score": score}
            predictions.append(enriched)
            all_predictions.append({**enriched, "participant": participant["name"]})

        played = stats["exact"] + stats["trend"] + stats["miss"]
        effectiveness = round((stats["points"] / (played * 3)) * 100, 1) if played else 0
        enriched_participants.append(
            {
                "name": participant["name"],
                "file": participant["file"],
                "error": participant.get("error"),
                "stats": {**stats, "played": played, "effectiveness": effectiveness},
                "predictions": predictions,
            }
        )

    enriched_participants.sort(
        key=lambda item: (
            -item["stats"]["points"],
            -item["stats"]["exact"],
            -item["stats"]["trend"],
            item["name"].lower(),
        )
    )
    for index, participant in enumerate(enriched_participants, start=1):
        participant["rank"] = index

    return {
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "rules": {"exact": 3, "trend": 1, "miss": 0},
        "quinielaDir": str(quiniela_dir),
        "api": api_meta,
        "participants": enriched_participants,
        "fixtures": fixtures,
        "predictions": all_predictions,
    }


class DashboardHandler(SimpleHTTPRequestHandler):
    quiniela_dir = Path(os.environ.get("QUINIELA_DATA_DIR", DEFAULT_QUINIELA_DIR)).resolve()

    def translate_path(self, path: str) -> str:
        parsed = urllib.parse.urlparse(path)
        clean_path = parsed.path.lstrip("/") or "index.html"
        return str(STATIC_DIR / clean_path)

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/state":
            params = urllib.parse.parse_qs(parsed.query)
            refresh = params.get("refresh", ["0"])[0] in {"1", "true", "yes"}
            self.send_json(build_state(self.quiniela_dir, refresh=refresh))
            return
        if parsed.path == "/api/config":
            self.send_json(
                {
                    "uploadEnabled": bool(admin_token()),
                    "apiConfigured": True,
                    "apiSource": "espn",
                }
            )
            return
        if parsed.path == "/health":
            self.send_json({"ok": True})
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/upload":
            self.handle_upload()
            return
        self.send_json({"error": "Ruta no encontrada."}, status=HTTPStatus.NOT_FOUND)

    def handle_upload(self) -> None:
        token = admin_token()
        if not token:
            self.send_json({"error": "ADMIN_TOKEN no esta configurado en el servidor."}, status=HTTPStatus.FORBIDDEN)
            return

        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0 or length > MAX_UPLOAD_BYTES * 2:
            self.send_json({"error": "Archivo demasiado grande o vacio."}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            self.send_json({"error": "Solicitud invalida."}, status=HTTPStatus.BAD_REQUEST)
            return

        provided = str(payload.get("token", ""))
        if not hmac.compare_digest(provided, token):
            self.send_json({"error": "Token incorrecto."}, status=HTTPStatus.UNAUTHORIZED)
            return

        filename = safe_filename(payload.get("filename"))
        if not filename.lower().endswith(".xlsx"):
            self.send_json({"error": "Solo se aceptan archivos .xlsx."}, status=HTTPStatus.BAD_REQUEST)
            return

        try:
            content = base64.b64decode(str(payload.get("contentBase64", "")), validate=True)
        except (binascii.Error, ValueError):
            self.send_json({"error": "No pude decodificar el archivo."}, status=HTTPStatus.BAD_REQUEST)
            return

        if not content or len(content) > MAX_UPLOAD_BYTES:
            self.send_json({"error": "El archivo excede el limite de 12 MB."}, status=HTTPStatus.BAD_REQUEST)
            return

        self.quiniela_dir.mkdir(parents=True, exist_ok=True)
        output_path = self.quiniela_dir / filename
        output_path.write_bytes(content)
        self.send_json({"ok": True, "filename": filename, "bytes": len(content)})

    def send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def admin_token() -> str:
    return os.environ.get("ADMIN_TOKEN", "").strip()


def safe_filename(value: Any) -> str:
    name = Path(str(value or "quiniela.xlsx")).name.strip()
    name = re.sub(r"[^\w .()-]+", "_", name, flags=re.UNICODE)
    name = re.sub(r"\s+", " ", name).strip(" .")
    if not name:
        name = "quiniela.xlsx"
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name


def main() -> None:
    ensure_output_streams()
    parser = argparse.ArgumentParser(description="Dashboard local para la Quiniela Mundial 2026.")
    parser.add_argument("--host", default=os.environ.get("HOST", "127.0.0.1"))
    parser.add_argument("--port", default=int(os.environ.get("PORT", "8765")), type=int)
    parser.add_argument("--quiniela-dir", default=os.environ.get("QUINIELA_DATA_DIR", str(DEFAULT_QUINIELA_DIR)))
    args = parser.parse_args()

    DashboardHandler.quiniela_dir = Path(args.quiniela_dir).resolve()
    server = ThreadingHTTPServer((args.host, args.port), DashboardHandler)
    print(f"Dashboard: http://{args.host}:{args.port}")
    print(f"Leyendo quinielas desde: {DashboardHandler.quiniela_dir}")
    print("Resultados desde ESPN public scoreboard.")
    server.serve_forever()


if __name__ == "__main__":
    main()
