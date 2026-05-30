from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl

from app import CSV_FIELDS, as_int, excel_date, excel_time, find_prediction_rows, participant_from_workbook


APP_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = APP_DIR.parent
DEFAULT_OUTPUT = APP_DIR / "data" / "quinielas.csv"


def export_excel_to_csv(input_dir: Path, output_path: Path) -> int:
    rows: list[dict[str, object]] = []
    files = sorted(path for path in input_dir.glob("*.xlsx") if not path.name.startswith("~$"))

    for path in files:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
        for worksheet in workbook.worksheets:
            prediction_rows = find_prediction_rows(worksheet)
            if not prediction_rows:
                continue

            participant = participant_from_workbook(path, worksheet)
            if len(workbook.worksheets) > 1 and worksheet.title.lower() != "quiniela":
                participant = f"{participant} ({worksheet.title})"

            for row in prediction_rows:
                rows.append(
                    {
                        "participant": participant,
                        "file": path.name,
                        "sheet": worksheet.title,
                        "matchNo": as_int(worksheet.cell(row, 5).value),
                        "date": excel_date(worksheet.cell(row, 6).value) or "",
                        "time": excel_time(worksheet.cell(row, 7).value) or "",
                        "home": worksheet.cell(row, 9).value or "",
                        "predHome": as_int(worksheet.cell(row, 10).value),
                        "predAway": as_int(worksheet.cell(row, 12).value),
                        "away": worksheet.cell(row, 13).value or "",
                        "venue": worksheet.cell(row, 15).value or "",
                    }
                )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convierte quinielas .xlsx a CSV para publicar el dashboard.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR), help="Carpeta donde estan los Excel.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="CSV de salida.")
    args = parser.parse_args()

    count = export_excel_to_csv(Path(args.input_dir), Path(args.output))
    print(f"CSV generado: {args.output}")
    print(f"Predicciones exportadas: {count}")


if __name__ == "__main__":
    main()
